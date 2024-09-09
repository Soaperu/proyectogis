#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Importar librerias
import arcpy
import os
import sys
import json
import pandas as pd
import csv
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles import Border, Side


def str_to_num(val):
    try:
        if '.' in val:
            num = float(val)
        else:
            num = int(val)
    except:
        num = val
    return num

def get_styled_xls(file_csv,file_xlsx):
    reload(sys)   
    sys.setdefaultencoding('utf8')
    wb = openpyxl.Workbook()
    ws = wb.active
    # definicion de colores, estilos de bordes y estilos
    color = "bfbfbf"
    color_resaltado = "ffc7ce"
    color_texto_resaltado = "9b0005"
    header_font = Font(name = 'Arial Narrow', bold=True)
    document_font = Font(name = 'Arial Narrow',size=10)
    font_resaltado = Font(color=color_texto_resaltado)
    formula_resaltado = "ABS(FF1)>10"
    fill_gray = PatternFill(start_color=color, end_color=color, fill_type="solid")
    fill_resaltado = PatternFill(start_color=color_resaltado, end_color=color_resaltado, fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))



    number_rows = 0
    # with open(file_csv ,encoding="utf-8-sig") as f:
    with open(file_csv) as f:
        reader = csv.reader(f, delimiter=';', escapechar='\n')
        for row in reader:
            row = [str_to_num(x) for x in row]
            ws.append(row)
            number_rows+=1

    # Creamos el diccionario de ancho de columna
    counter_row = 0
    dims = {}
    for row in ws.rows:
        for cell in row:
            col = column_index_from_string(cell.column)
            cell.font = document_font

            cell.border = thin_border
            if counter_row==0:
                dims[cell.column] =len(str(cell.value))+3
                cell.fill= fill_gray
                cell.font = header_font
            if (ws.cell(row=1, column=col).value).lower() in [u'descripción litológica', 'observaciones']:
                dims[cell.column] = 50
            else:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))
        counter_row+=1

    # Aplicamos la dimension de ancho de columnas
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
    # ws.row_dimensions[1].heigth = 10


    # Le aplicamos formato de numero a las columnas de meq
    for row in ws:
        for cell in row:
            col = column_index_from_string(cell.column)
            cellvalue = (ws.cell(row=1, column=col).value).lower()
            if ('meq' in cellvalue) or (u'%' in cellvalue) :
                cell.number_format = '0.000'
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Aplicando regla para los mayores a 10
    ws.conditional_formatting.add('FF1:FF{}'.format(number_rows),FormulaRule(formula=[formula_resaltado], stopIfTrue=True, fill=fill_resaltado, font = font_resaltado))
    wb.save(file_xlsx)
    os.remove(file_csv)

if __name__ == '__main__':
    csvfile = sys.argv[1]
    xls = sys.argv[2]
    get_styled_xls(csvfile,xls)